from openpyxl import load_workbook

wb = load_workbook("transactions.xlsx")
ws = wb['sheet_1']

ws['D1'] = "corrected"

for row in range(2, ws.max_row + 1):
    
    price = ws.cell(row=row, column=3).value
    
    if price is not None:
        
        corrected_price = price * 0.9
        ws.cell(row=row, column=4, value=corrected_price)
        ws.cell(row=row, column=4).number_format = '#,##0.00" €"'

wb.save("transactions.xlsx")

wb.close()
